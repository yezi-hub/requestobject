import openpyxl
import os

from openpyxl import load_workbook
import time
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill
import traceback

class Excel:
    #构造方法，传入excel文件的绝对路径
    def __init__(self, excel_file_path=None):
        if not excel_file_path:
            self.wb = None
            self.sheet = None
            return
        if os.path.exists(excel_file_path):  # 判断excel文件的路径是否存在
            self.excel_file_path = excel_file_path  # 存到实例变量中
            # 把excel加载到内存中
            self.wb = load_workbook(self.excel_file_path)
        else:
            print("%s 的文件不存在，请重新设定" % excel_file_path)
            self.wb = None
        self.sheet = None  # 初始化sheet的内容为None

    #获取excel的文件路径
    def get_file_path(self):  # 获取excel的绝对路径
        return self.excel_file_path

    #设置excel的文件路径
    def set_file_path(self, excel_file_path):  # 设定新的excel文件路径
        if os.path.exists(excel_file_path):  # 判断excel文件的路径是否存在
            self.excel_file_path = excel_file_path  # 存到实例变量中
            self.wb = load_workbook(self.excel_file_path)#根据设定的路径，设定wb对象
            self.sheet = None
        else:
            print("%s 的文件不存在，请重新设定" % excel_file_path)

    #获取excel文件中的所有sheet名称
    def get_sheet_names(self):  # 获取所有的sheet名称
        if self.wb:
            return self.wb.sheetnames
        else:
            return []

    #设置当前操作的sheet对象
    def set_sheet(self, sheet_name):  # 设定我要操作哪个sheet对象
        print("#########get_sheet_names(): ", self.get_sheet_names())
        if sheet_name in self.get_sheet_names():  # 判断是否是个有效的sheet名称
            self.sheet = self.wb[sheet_name]
        else:
            print("设定的sheet名称 %s 不存在！" % sheet_name)

    def create_sheet(self,sheet_name):
        if sheet_name in self.get_sheet_names():
            print("sheet %s 已经存在无需创建！" %sheet_name)
            return False
        if self.wb:
            try:
                self.wb.create_sheet(sheet_name)
                return True
            except Exception as e:
                print("sheet %s 创建时出现异常，异常信息：%s！" %(sheet_name,e))
                return False
        else:
            print("当前没有设定excel文件的路径，无法新建任何sheet！")
            return False

    #获取当前操作的sheet对象
    def get_current_sheet_name(self):  # 我当前正在操作的sheet对象的名字
        if self.sheet:
            return self.sheet.title
        else:
            print("没有设定当前操作的sheet!")

    #获取所有的行对象
    def get_all_rows(self):  # 获取当前sheet中的所有行对象
        if not self.sheet:  # 如果没有设定当前操作的sheet，就返回空
            return []

        rows = []
        for row in self.sheet.iter_rows():
            rows.append(row)

        return rows

    #获取所有行行中的值，返回的是个二维列表
    def get_all_rows_values(self):  # 获取当前sheet中的所有行对象
        if not self.get_all_rows():  # 如果没有数据，就返回[]
            return []
        all_rows_values = []
        for row in self.get_all_rows():  # 遍历每一行的对象
            row_values = []
            for cell in row:  # 遍历每一行中的单元格对象
                row_values.append(cell.value)  # 把单元格的值存储到row_values列表中

            all_rows_values.append(row_values)
        return all_rows_values  # 返回的结果是一个二维列表

    #根据行号获取行对象
    def get_row_by_row_no(self, row_no):  # 获取行号对应的行对象，行号从1开始
        if not isinstance(row_no, int):
            print("输入的行号参数 %s 不是整数，请重新设定！" % row_no)
            return []

        all_rows = self.get_all_rows()
        if not all_rows:  # sheet为空
            return []
        if row_no < 1 or row_no > len(all_rows):  # 行号非法的情况
            print("行号小于1，或者行号大于了有效数据行的范围！")
            return []

        return all_rows[row_no - 1]

    #根据行号，获取行中的所有值
    def get_row_value_by_row_no(self, row_no):  # 取出某一行的值
        if not isinstance(row_no, int):
            print("输入的行号参数 %s 不是整数，请重新设定！" % row_no)
            return []

        if not self.get_row_by_row_no(row_no):
            return []

        row = self.get_row_by_row_no(row_no)
        row_values = []
        for cell in row:
            row_values.append(cell.value)

        return row_values

    #根据列号获取列对象
    def get_col_by_col_no(self, col_no):  # 获得某一列的对象
        if not isinstance(col_no, int):
            print("输入的列号参数 %s 不是整数，请重新设定！" % col_no)
            return []

        if not self.get_all_rows():
            return []

        cols = []  # 用于存储所有的列，每一列是一个元组，cols是一个二维列表
        for col in self.sheet.iter_cols():
            cols.append(col)

        max_col = len(cols)
        if col_no < 1 or col_no > max_col:
            print("列号%s不是有效范围内的列表！" % col_no)
            return []

        return cols[col_no - 1]

    #根据列号，获得某一列的值
    def get_col_value_by_col_no(self, col_no):  # 获得某一列的值
        if not isinstance(col_no, int):
            print("输入的列号参数 %s 不是整数，请重新设定！" % col_no)
            return []

        col = self.get_col_by_col_no(col_no)
        if not col:
            return []

        values = []
        for cell in col:
            values.append(cell.value)

        return values

    #根据行号和列号，获得某个单元格对象
    def get_cell(self, row_no, col_no):  # 读取某个单元格对象的方法

        if not isinstance(row_no, int):
            print("输入的行号参数 %s 不是整数，请重新设定！" % row_no)
            return None

        if not isinstance(col_no, int):
            print("输入的列号参数 %s 不是整数，请重新设定！" % col_no)
            return None

        try:
            cell = self.sheet.cell(row=row_no, column=col_no)
            return cell
        except Exception as e:
            print(e)
            print("读取 %s 行 %s 列的单元格对象失败" % (row_no, col_no))

    #根据行号和列号，获取某个单元格中的值，行号从1开始
    def get_cell_value(self, row_no, col_no):  # 读取某个单元格的值的方法
        if not isinstance(row_no, int):
            print("输入的行号参数 %s 不是整数，请重新设定！" % row_no)
            return None

        if not isinstance(col_no, int):
            print("输入的列号参数 %s 不是整数，请重新设定！" % col_no)
            return None

        try:
            value = self.get_cell(row_no, col_no).value
            return value
        except Exception as e:
            print(e)
            return None

    #根据行号和列表，写入单元格的值
    def write_cell(self, row_no, col_no, value):
        if not isinstance(row_no, int):
            print("输入的行号参数 %s 不是整数，请重新设定！" % row_no)
            return None

        if not isinstance(col_no, int):
            print("输入的列号参数 %s 不是整数，请重新设定！" % col_no)
            return None

        try:
            self.get_cell(row_no, col_no).value = value
            self.save()
        except Exception as e:
            print(e)
            print("写入单元格%s行%列的值%s失败了！" % (row_no, col_no, value))

    #根据行号和列号，在单元格中=写入日期和时间
    def write_cell_date_time(self,row_no,col_no):
        if not isinstance(row_no, int):
            print("输入的行号参数 %s 不是整数，请重新设定！" % row_no)
            return None

        if not isinstance(col_no, int):
            print("输入的列号参数 %s 不是整数，请重新设定！" % col_no)
            return None
        date_time = time.strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.get_cell(row_no, col_no).value = date_time
            self.save()
        except Exception as e:
            print(e)
            print("写入单元格%s行%s列的值%s失败了！" % (row_no, col_no,date_time))

    #在sheet中的最后面，追加写入一行
    def write_a_line(self,line,fill=None):
        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        if fill:#把fill的参数转变为可以使用的填充对象
            if fill.lower()=="blue":
                fill = PatternFill(start_color='00CCFF', end_color='00CCFF', fill_type='solid')
            elif fill.lower()=="green":
                fill = PatternFill(fill_type="solid",    start_color='00FF00',    end_color='00FF00')
            elif fill.lower() == "red":
                fill = PatternFill(fill_type="solid", start_color='FF0000', end_color='FF0000')
            elif fill.lower() == "yellow":
                fill = PatternFill(fill_type="solid", start_color='FFFF00', end_color='FFFF00')
            elif fill.lower() == "wathet":
                fill = PatternFill(fill_type="solid", start_color='33CCFF', end_color='33CCFF')
            else:
                fill = PatternFill(fill_type="solid", start_color='FFFFFF', end_color='FFFFFF')

        if self.wb and self.sheet and line:
            try:
                self.sheet.append(line)
                for cell in list(self.sheet.rows)[-1]:
                    cell.border = border#添加边框
                    if fill:
                        cell.fill = fill#添加背景色
                    if "失败" in str(cell.value) or "fail" in str(cell.value).lower():
                        cell.fill = PatternFill(fill_type="solid", start_color='FF0000', end_color='FF0000')
                return True
            except Exception as e:
                traceback.print_exc()
                print("写入行内容 %s 失败，失败原因：%s" %(line,e))
                return False
        else:
            traceback.print_exc()
            print("写入行内容 %s 失败，失败原因：excel文件没有设定，或者sheet没有设定" % (line))
            return False

    #写入多行，参数是一个二维列表
    def write_lines(self,rows,header_color=None):
        try:
            for row_no in range(len(rows)):
                if row_no==0:
                    self.write_a_line(rows[row_no],header_color)
                else:
                    self.write_a_line(rows[row_no])
        except Exception as e:
            print("写入多行内容%s时候，出现异常，信息：%s" %(rows,e))


    #保存文件
    def save(self):
        if self.wb:
            self.wb.save(self.get_file_path())
        else:
            print("当前excel文件没有设置路径！无法保存文件内容")

def validate_excel_and_sheet(test_data_execel_wb, test_step_sheet_name):
    if not test_data_execel_wb:
        print("要操作的excel文件对象不存在！")
        raise Exception("要操作的excel文件对象不存在！")

    # sheet名称在exel文件中是否存在
    if test_step_sheet_name not in test_data_execel_wb.get_sheet_names():
        print("sheet名称:%s 在excel文件:%s 中不存在！" % (test_step_sheet_name, test_data_execel_wb.get_file_path()))
        raise Exception(
            "sheet名称:%s 在excel文件:%s 中不存在！" % (test_step_sheet_name, test_data_execel_wb.get_file_path()))

if __name__ == "__main__":
    excel = Excel("e:\\a.xlsx")
    # print(excel.get_file_path())
    # excel.set_file_path("e:\\b.xlsx")
    # print(excel.get_file_path())

    # print(excel.get_sheet_names())
    # print(excel.get_current_sheet_name())
    excel.set_sheet("Sheet1")
    # print(excel.get_current_sheet_name())
    # excel.set_sheet("Sheet2")
    # print(excel.get_all_rows())
    # print(excel.get_row_by_row_no(1))
    # print(excel.get_row_by_row_no(3))
    # print(excel.get_all_rows_values())
    # print(excel.get_row_value_by_row_no(3))
    # print(excel.get_col_by_col_no(1))
    # print(excel.get_col_by_col_no(3))
    # print(excel.get_col_value_by_col_no(2))
    # print(excel.get_cell(1,1))
    # print(excel.get_cell_value(1,1))
    # excel.save()
    #excel.write_cell(1, 1, 200)
    #excel1 = Excel()
    #excel1.create_sheet("测试")
    #excel1.set_file_path("e:\\sample.xlsx")
    #excel1.create_sheet("测试")
    #excel1.save()

    #excel.write_cell_date_time(5,5)
    #excel.set_sheet("Sheet1")
    #excel.write_a_line(["今天","天气","不错",5,6,7])
    #excel.write_a_line(["今天", "天气", "不错", 5, 6, 7],"green")
    #excel.write_a_line(["今天", "天气", "失败", 5, 6, 7], "blue")
    #excel.write_a_line(["今天", "天气", "Fail", 5, 6, 7], "yellow")
    #excel.save()

    excel.set_sheet("Sheet1")
    excel.write_lines([["今天", "中秋节", 1, 2], ["明天", "上班", 3, 4], [5, 6, 7, 8]])
    excel.write_lines([["今天","中秋节",1,2],["明天","上班",3,4],[5,6,7,8]],"green")
    excel.save()


    #excel1 = Excel()
    #excel1.write_a_line(["今天", "天气", "不错", 1, 2, 4])

    #excel1 = Excel("e:\\a.xlsx")
    #excel1.write_a_line(["今天", "天气", "不错", 1, 2, 4])
    #excel1.save()
